"""Bank statement file parsers — CSV / Excel / PDF text. No bank credentials."""

from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Protocol

from app.modules.banks.parsers import NormalizedBankTx


class StatementParser(Protocol):
    def parse(self, *, filename: str, content: bytes, bank_code: str) -> list[NormalizedBankTx]: ...


_AMOUNT_RE = re.compile(r"(?P<amount>-?\d{1,3}(?:[\s\u00a0]\d{3})*(?:[.,]\d{2})|-?\d+[.,]\d{2})")
_DATE_RE = re.compile(r"(?P<d>\d{2}[./]\d{2}[./]\d{2,4}|\d{4}-\d{2}-\d{2})")


def _parse_amount(raw: str) -> Decimal | None:
    cleaned = (
        raw.strip()
        .replace("\u00a0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("RUB", "")
        .replace("руб.", "")
        .replace("руб", "")
    )
    cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(raw: str) -> datetime | None:
    text = raw.strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def _row_tx(
    *,
    bank_code: str,
    idx: int,
    amount: Decimal,
    merchant: str,
    booked_at: datetime,
) -> NormalizedBankTx:
    digest = hashlib.sha256(
        f"{bank_code}|{booked_at.isoformat()}|{amount}|{merchant}|{idx}".encode()
    ).hexdigest()[:24]
    direction = "credit" if amount < 0 else "debit"
    return NormalizedBankTx(
        external_id=f"stmt-{bank_code}-{digest}",
        amount=abs(amount),
        currency="RUB",
        merchant_raw=merchant[:255] or "Операция",
        booked_at=booked_at,
        direction=direction,
    )


class CsvStatementParser:
    def parse(self, *, filename: str, content: bytes, bank_code: str) -> list[NormalizedBankTx]:
        text = content.decode("utf-8-sig", errors="replace")
        # Try delimiter detection
        sample = text[:2048]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames is None:
            return self._parse_loose_lines(text, bank_code)

        fields = {f.lower().strip(): f for f in reader.fieldnames if f}
        date_key = _pick(fields, ("дата", "date", "booking date", "операция"))
        amount_key = _pick(fields, ("сумма", "amount", "sum", "оборот"))
        merchant_key = _pick(
            fields, ("описание", "merchant", "назначение", "details", "контрагент", "place")
        )
        if not amount_key:
            return self._parse_loose_lines(text, bank_code)

        txs: list[NormalizedBankTx] = []
        for i, row in enumerate(reader):
            amount = _parse_amount(str(row.get(amount_key, "")))
            if amount is None or amount == 0:
                continue
            booked = (
                _parse_date(str(row.get(date_key, ""))) if date_key else None
            ) or datetime.now(UTC)
            merchant = str(row.get(merchant_key, "")).strip() if merchant_key else "Операция"
            txs.append(
                _row_tx(
                    bank_code=bank_code,
                    idx=i,
                    amount=amount,
                    merchant=merchant,
                    booked_at=booked,
                )
            )
        return txs

    def _parse_loose_lines(self, text: str, bank_code: str) -> list[NormalizedBankTx]:
        txs: list[NormalizedBankTx] = []
        for i, line in enumerate(text.splitlines()):
            am = _AMOUNT_RE.search(line)
            if not am:
                continue
            amount = _parse_amount(am.group("amount"))
            if amount is None:
                continue
            dm = _DATE_RE.search(line)
            booked = _parse_date(dm.group("d")) if dm else datetime.now(UTC)
            merchant = _AMOUNT_RE.sub("", line)
            if dm:
                merchant = merchant.replace(dm.group("d"), "")
            merchant = re.sub(r"\s+", " ", merchant).strip(" ;|-") or "Операция"
            txs.append(
                _row_tx(
                    bank_code=bank_code, idx=i, amount=amount, merchant=merchant, booked_at=booked
                )
            )
        return txs


class ExcelStatementParser:
    """XLSX via openpyxl if installed; otherwise treat as unsupported."""

    def parse(self, *, filename: str, content: bytes, bank_code: str) -> list[NormalizedBankTx]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Для Excel установите openpyxl или загрузите CSV") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Convert to CSV-like text for reuse
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        for row in rows:
            writer.writerow(["" if c is None else str(c) for c in row])
        return CsvStatementParser().parse(
            filename=filename, content=buf.getvalue().encode("utf-8"), bank_code=bank_code
        )


class PdfStatementParser:
    def parse(self, *, filename: str, content: bytes, bank_code: str) -> list[NormalizedBankTx]:
        text = ""
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(content))
            parts = []
            for page in reader.pages:
                parts.append(page.extract_text() or "")
            text = "\n".join(parts)
        except Exception:
            # Best-effort: latin1 dump of binary is useless; fail clearly
            raise ValueError(
                "Не удалось извлечь текст из PDF — сохраните выписку как CSV"
            ) from None
        if not text.strip():
            raise ValueError("PDF без текста — экспортируйте выписку в CSV")
        return CsvStatementParser()._parse_loose_lines(text, bank_code)


def _pick(fields: dict[str, str], candidates: tuple[str, ...]) -> str | None:
    for c in candidates:
        if c in fields:
            return fields[c]
        for key, original in fields.items():
            if c in key:
                return original
    return None


def parse_statement_file(
    *, filename: str, content: bytes, bank_code: str
) -> list[NormalizedBankTx]:
    lower = filename.lower()
    code = bank_code.lower()
    if code in {"t-bank", "tbank", "тинькофф"}:
        code = "tinkoff"
    if code in {"сбер", "sberbank"}:
        code = "sber"
    if code in {"альфа", "alfabank"}:
        code = "alfa"
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return CsvStatementParser().parse(filename=filename, content=content, bank_code=code)
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return ExcelStatementParser().parse(filename=filename, content=content, bank_code=code)
    if lower.endswith(".pdf"):
        return PdfStatementParser().parse(filename=filename, content=content, bank_code=code)
    # sniff
    if content[:4] == b"%PDF":
        return PdfStatementParser().parse(filename=filename, content=content, bank_code=code)
    return CsvStatementParser().parse(filename=filename, content=content, bank_code=code)
